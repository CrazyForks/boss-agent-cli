"""Incremental JSON, CSV and formatted XLSX output for crawl runs."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

FIELDS = (
	"query", "page", "rank", "title", "salary", "city", "district", "business_district",
	"company", "company_scale", "industry", "education", "experience", "labels", "benefits", "post_description",
	"address", "detail_status",
)


def write_run_outputs(output_dir: Path, rows: list[dict[str, Any]]) -> dict[str, str]:
	"""Overwrite the three deterministic artifacts after each completed page."""
	export_rows = [{field: row.get(field, "") for field in FIELDS} for row in rows]
	output_dir.mkdir(parents=True, exist_ok=True)
	json_path = output_dir / "jobs.json"
	csv_path = output_dir / "jobs.csv"
	xlsx_path = output_dir / "jobs.xlsx"
	json_path.write_text(json.dumps(export_rows, ensure_ascii=False, indent=2), encoding="utf-8")
	with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
		writer = csv.DictWriter(file, fieldnames=FIELDS, extrasaction="ignore")
		writer.writeheader()
		writer.writerows(export_rows)
	_write_xlsx(xlsx_path, export_rows)
	return {"json": str(json_path), "csv": str(csv_path), "xlsx": str(xlsx_path)}


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Alignment
		from openpyxl.worksheet.table import Table, TableStyleInfo
		from openpyxl.utils import get_column_letter
	except ImportError as exc:  # pragma: no cover - exercised through CLI dependency error
		raise RuntimeError("crawl XLSX 导出需要安装 boss-agent-cli[crawl]") from exc

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "jobs"
	for index, field in enumerate(FIELDS, 1):
		cell = sheet.cell(row=1, column=index, value=field)
		cell.alignment = Alignment(horizontal="center", vertical="center")
	for row in rows:
		sheet.append([row.get(field, "") for field in FIELDS])

	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{max(1, len(rows) + 1)}"
	sheet.row_dimensions[1].height = 22
	for index, field in enumerate(FIELDS, 1):
		letter = get_column_letter(index)
		width = 18
		if field == "post_description":
			width = 54
		elif field in {"labels", "benefits", "address"}:
			width = 28
		sheet.column_dimensions[letter].width = width
	for row_index in range(2, len(rows) + 2):
		sheet.row_dimensions[row_index].height = 20
		for cell in sheet[row_index]:
			# 保留完整值，但在表格中固定为单行，长内容不拉高整行。
			cell.alignment = Alignment(vertical="center", wrap_text=False)

	if rows:
		table = Table(displayName="JobsTable", ref=f"A1:{get_column_letter(len(FIELDS))}{len(rows) + 1}")
		table.tableStyleInfo = TableStyleInfo(
			name="TableStyleMedium2",
			showFirstColumn=False,
			showLastColumn=False,
			showRowStripes=True,
			showColumnStripes=False,
		)
		sheet.add_table(table)
	workbook.save(path)
