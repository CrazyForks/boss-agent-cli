import json
import sqlite3
import sys
import types
from hashlib import sha256
from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from boss_agent_cli.cache.store import CacheStore
from boss_agent_cli.config import DEFAULTS
from boss_agent_cli.crawler.exporter import write_run_outputs
from boss_agent_cli.crawler.hooks import HOOK_SCRIPT_NAMES, HookInjection, HookRegistrationError, inject_hook_profile
from boss_agent_cli.crawler.operations import crawl_results, import_crawl_shortlist
from boss_agent_cli.crawler.service import CrawlBudget, CrawlService, CrawlSettings, CrawlStopRequested
from boss_agent_cli.crawler.transport import JOBLIST_TARGET, DrissionCrawlerSession
from boss_agent_cli.main import cli


def _job(job_id: str, security_id: str) -> dict:
	return {
		"encryptJobId": job_id,
		"securityId": security_id,
		"jobName": "AI 工程师",
		"salaryDesc": "20-30K",
		"cityName": "杭州",
		"brandName": "测试公司",
		"jobLabels": ["Python", "LLM"],
		"welfareList": ["双休"],
	}


def _page(*jobs: dict, has_more: bool = False, code: int = 0) -> dict:
	return {"code": code, "zpData": {"jobList": list(jobs), "hasMore": has_more}}


def _write_hooks(root: Path) -> Path:
	root.mkdir(parents=True, exist_ok=True)
	lines = []
	for index, (_, filename) in enumerate(HOOK_SCRIPT_NAMES, 1):
		payload = f"globalThis.hook{index} = true;\n".encode()
		(root / filename).write_bytes(payload)
		lines.append(f"{filename} {sha256(payload).hexdigest()}")
	(root / "SHA256SUMS").write_text("\n".join(lines) + "\n", encoding="ascii")
	return root


class _NoDelayBudget(CrawlBudget):
	def wait(self, kind: str) -> None:
		self._cache.put_crawl_budget(f"test:{kind}", 1.0)


class _FakeTransport:
	def __init__(
		self,
		pages: dict[int, dict] | None = None,
		*,
		fail_pages: set[int] | None = None,
		detail_payloads: dict[str, dict] | None = None,
		stop_on_detail: bool = False,
	) -> None:
		self.pages = pages or {}
		self.fail_pages = fail_pages or set()
		self.detail_payloads = detail_payloads or {}
		self.stop_on_detail = stop_on_detail
		self.page_calls: list[int] = []
		self.detail_calls: list[str] = []
		self.closed = False
		self.hooks: list[HookInjection] = []

	def open(self) -> list[HookInjection]:
		return self.hooks

	def fetch_page(self, query: str, city_code: str, page_no: int) -> dict:
		self.page_calls.append(page_no)
		if page_no in self.fail_pages:
			raise TimeoutError("listener timeout")
		return self.pages[page_no]

	def fetch_detail(self, security_id: str) -> dict:
		self.detail_calls.append(security_id)
		if self.stop_on_detail:
			raise CrawlStopRequested("stop requested")
		return self.detail_payloads.get(
			security_id,
			{"code": 0, "zpData": {"jobCard": {"postDescription": "职位描述", "address": "西湖区"}}},
		)

	def close(self) -> None:
		self.closed = True


def _settings(
	tmp_path: Path,
	*,
	pages: int = 5,
	with_detail: bool = False,
	max_requests: int = 20,
	max_details: int = 50,
	max_seconds: int = 600,
	max_retries: int = 1,
) -> CrawlSettings:
	return CrawlSettings(
		query="AI",
		city_code="101210100",
		pages=pages,
		with_detail=with_detail,
		profile_path=tmp_path / "profile",
		chrome_path=None,
		cdp_port=9222,
		hook_profile="none",
		max_requests=max_requests,
		max_details=max_details,
		max_seconds=max_seconds,
		max_retries=max_retries,
		operating_mode="research",
	)


def _service(tmp_path: Path, transport_factory):
	cache = CacheStore(tmp_path / "cache.db")
	return cache, CrawlService(
		cache,
		data_dir=tmp_path,
		transport_factory=transport_factory,
		budget_factory=lambda store: _NoDelayBudget(store),
	)


def _enable_research(runner: CliRunner, tmp_path: Path) -> None:
	result = runner.invoke(cli, ["--data-dir", str(tmp_path), "--json", "config", "set", "operating_mode", "research"])
	assert result.exit_code == 0, result.output


def test_screenshot_full_hook_registers_exactly_seven_user_scripts(tmp_path):
	class Page:
		def __init__(self) -> None:
			self.calls: list[tuple[str, dict]] = []

		def run_cdp(self, command: str, **kwargs) -> None:
			self.calls.append((command, kwargs))

	page = Page()
	results = inject_hook_profile(page, "screenshot-full", _write_hooks(tmp_path / "hooks"))
	assert [item.name for item in results] == [name for name, _ in HOOK_SCRIPT_NAMES]
	assert len(results) == 7
	assert all(item.success and item.sha256 for item in results)
	assert all(command == "Page.addScriptToEvaluateOnNewDocument" for command, _ in page.calls)


def test_hook_requires_manifest_and_rejects_modified_file(tmp_path):
	root = _write_hooks(tmp_path / "hooks")
	(root / "hook_clear.js").write_text("changed", encoding="utf-8")
	try:
		inject_hook_profile(object(), "screenshot-full", root)
	except HookRegistrationError as exc:
		failed = next(item for item in exc.injections if item.name == "hook_clear")
		assert "SHA-256" in failed.reason
	else:
		raise AssertionError("expected HookRegistrationError")

	(root / "SHA256SUMS").unlink()
	try:
		inject_hook_profile(object(), "screenshot-full", root)
	except ValueError as exc:
		assert "SHA256SUMS" in str(exc)
	else:
		raise AssertionError("expected ValueError")


def test_joblist_listener_target_matches_reference_endpoint():
	assert JOBLIST_TARGET == r"wapi/zpgeek/search/joblist\.json"


def test_hook_registration_happens_before_first_navigation(monkeypatch, tmp_path):
	events: list[str] = []

	class Options:
		def set_local_port(self, port: int) -> None:
			pass

		def set_user_data_path(self, path: str) -> None:
			pass

	class Listener:
		def start(self, *args, **kwargs) -> None:
			events.append("listen")

		def wait(self, **kwargs):
			return types.SimpleNamespace(response=types.SimpleNamespace(body=_page()))

	class Page:
		url = ""
		html = ""

		def __init__(self) -> None:
			self.listen = Listener()

		def run_cdp(self, command: str, **kwargs) -> None:
			events.append("hook")

		def get(self, url: str) -> None:
			events.append("navigate")

		def quit(self) -> None:
			pass

	page = Page()
	monkeypatch.setitem(sys.modules, "DrissionPage", types.SimpleNamespace(ChromiumOptions=Options, ChromiumPage=lambda options: page))
	session = DrissionCrawlerSession(
		profile_path=tmp_path / "profile",
		chrome_path=None,
		cdp_port=9222,
		hook_profile="screenshot-full",
		hook_dir=_write_hooks(tmp_path / "hooks"),
	)
	session.open()
	session.fetch_page("AI", "101210100", 1)
	assert events.count("hook") == 7
	assert max(index for index, event in enumerate(events) if event == "hook") < events.index("navigate")


def test_listener_parses_string_json_response(monkeypatch, tmp_path):
	class Options:
		def set_local_port(self, port: int) -> None:
			pass

		def set_user_data_path(self, path: str) -> None:
			pass

	class Listener:
		def start(self, *args, **kwargs) -> None:
			pass

		def wait(self, **kwargs):
			return types.SimpleNamespace(response=types.SimpleNamespace(body=json.dumps(_page(_job("job-1", "sec-1")))))

	class Page:
		url = ""
		html = ""
		listen = Listener()

		def get(self, url: str) -> None:
			pass

		def quit(self) -> None:
			pass

	page = Page()
	monkeypatch.setitem(sys.modules, "DrissionPage", types.SimpleNamespace(ChromiumOptions=Options, ChromiumPage=lambda options: page))
	session = DrissionCrawlerSession(
		profile_path=tmp_path / "profile", chrome_path=None, cdp_port=9222, hook_profile="none", hook_dir=None,
	)
	session.open()
	assert session.fetch_page("AI", "101210100", 1)["zpData"]["jobList"][0]["encryptJobId"] == "job-1"


def test_crawl_writes_outputs_and_reuses_cached_detail(tmp_path):
	transport = _FakeTransport({1: _page(_job("job-1", "sec-1"))})
	cache, service = _service(tmp_path, lambda settings: transport)
	cache.put_job_desc("job-1", "缓存职位描述")
	outcome = service.create_and_run(_settings(tmp_path, with_detail=True))
	assert outcome.status == "completed"
	assert transport.detail_calls == []
	assert all(Path(path).exists() for path in outcome.output_paths.values())
	workbook = load_workbook(outcome.output_paths["xlsx"])
	sheet = workbook["jobs"]
	assert sheet.freeze_panes == "A2"
	assert sheet.auto_filter.ref == "A1:R2"
	assert sheet["P2"].alignment.wrap_text is not True
	assert sheet.row_dimensions[2].height == 20


def test_crawl_honors_page_limit_and_resume_deduplicates(tmp_path):
	first = _FakeTransport({1: _page(_job("job-1", "sec-1"), has_more=True), 2: _page(code=37)})
	second = _FakeTransport({2: _page(_job("job-2", "sec-2"), has_more=False)})
	transports = iter((first, second))
	cache, service = _service(tmp_path, lambda settings: next(transports))
	stopped = service.create_and_run(_settings(tmp_path, pages=3))
	assert stopped.status == "risk_stopped"
	assert stopped.next_page == 2
	resumed = service.resume(stopped.run_id, operating_mode="research")
	assert resumed.status == "completed"
	assert [item["job_key"] for item in cache.list_crawl_jobs(stopped.run_id)] == ["job-1", "job-2"]


def test_crawl_retries_once_then_checkpoints(tmp_path):
	transport = _FakeTransport(fail_pages={1})
	cache, service = _service(tmp_path, lambda settings: transport)
	outcome = service.create_and_run(_settings(tmp_path))
	assert outcome.status == "stopped"
	assert transport.page_calls == [1, 1]
	assert cache.get_crawl_run(outcome.run_id)["error"] == "listener timeout"


def test_budget_stops_and_persists_across_resume(tmp_path):
	first = _FakeTransport({1: _page(_job("job-1", "sec-1"), has_more=True)})
	second = _FakeTransport({2: _page(_job("job-2", "sec-2"), has_more=False)})
	transports = iter((first, second))
	cache, service = _service(tmp_path, lambda settings: next(transports))
	stopped = service.create_and_run(_settings(tmp_path, pages=2, max_requests=1))
	assert stopped.status == "budget_stopped"
	run = cache.get_crawl_run(stopped.run_id)
	assert run["requests_attempted"] == 1
	resumed = service.resume(stopped.run_id, operating_mode="research")
	assert resumed.status == "budget_stopped"
	assert second.page_calls == []


def test_stop_request_checkpoints_before_detail(tmp_path):
	transport = _FakeTransport({1: _page(_job("job-1", "sec-1"))})
	cache, service = _service(tmp_path, lambda settings: transport)
	run_id = service.create(_settings(tmp_path, with_detail=True))
	cache.request_crawl_stop(run_id)
	outcome = service.resume(run_id, operating_mode="research")
	assert outcome.status == "stopped"
	assert "stop requested" in outcome.error
	assert transport.page_calls == []


def test_public_results_redact_identifiers(tmp_path):
	cache = CacheStore(tmp_path / "cache.db")
	cache.create_crawl_run("run-1", _settings(tmp_path).as_dict(), str(tmp_path / "crawl" / "runs" / "run-1"))
	cache.put_crawl_job("run-1", "job-1", 1, {
		"job_id": "job-1", "security_id": "sec-1", "boss_name": "Boss", "title": "职位", "company": "公司",
	}, detail_done=False)
	job = crawl_results(cache, "run-1")["jobs"][0]
	assert "job_id" not in job and "security_id" not in job and "boss_name" not in job
	assert job["selector"].startswith("csel_")


def test_crawl_selector_import_is_stable_private_and_validated(tmp_path):
	cache = CacheStore(tmp_path / "cache.db")
	cache.create_crawl_run("run-1", _settings(tmp_path).as_dict(), str(tmp_path / "crawl" / "runs" / "run-1"))
	cache.put_crawl_job("run-1", "job-1", 1, {"job_id": "job-1", "security_id": "sec-1", "title": "职位"}, detail_done=False)
	selector = crawl_results(cache, "run-1")["jobs"][0]["selector"]
	assert selector == crawl_results(cache, "run-1")["jobs"][0]["selector"]

	imported = import_crawl_shortlist(cache, "run-1", selectors=(selector,))
	assert imported["imported"] == [{"selector": selector, "title": "职位"}]
	for invalid in (("missing",), (selector, selector)):
		try:
			import_crawl_shortlist(cache, "run-1", selectors=invalid)
		except ValueError:
			pass
		else:
			raise AssertionError("expected selector validation error")


def test_crawl_selector_migration_backfills_legacy_rows(tmp_path):
	db_path = tmp_path / "legacy.db"
	with sqlite3.connect(db_path) as connection:
		connection.execute(
			"CREATE TABLE crawl_jobs (run_id TEXT NOT NULL, job_key TEXT NOT NULL, page_no INTEGER NOT NULL, "
			"payload TEXT NOT NULL, detail_done INTEGER NOT NULL DEFAULT 0, updated_at REAL NOT NULL, "
			"PRIMARY KEY (run_id, job_key))"
		)
		connection.execute(
			"INSERT INTO crawl_jobs VALUES (?, ?, ?, ?, ?, ?)",
			("run-1", "job-1", 1, json.dumps({"job_id": "job-1"}), 0, 1.0),
		)
	with CacheStore(db_path) as cache:
		item = cache.list_crawl_jobs("run-1")[0]
		assert item["selector"].startswith("csel_")


def test_cli_rejects_assisted_mode_and_non_positive_pages(tmp_path, monkeypatch):
	monkeypatch.setitem(DEFAULTS, "operating_mode", "assisted")
	runner = CliRunner()
	assisted = runner.invoke(cli, ["--data-dir", str(tmp_path), "--json", "crawl", "run", "AI", "--city", "杭州"])
	assert assisted.exit_code == 1
	assert json.loads(assisted.output)["error"]["code"] == "COMPLIANCE_BLOCKED"
	_enable_research(runner, tmp_path)
	invalid_pages = runner.invoke(cli, ["--data-dir", str(tmp_path), "--json", "crawl", "run", "AI", "--city", "杭州", "--pages", "0"])
	assert invalid_pages.exit_code == 1


def test_crawl_start_and_stop_require_research_and_record_kill_switch(tmp_path, monkeypatch):
	launched: list[str] = []
	monkeypatch.setattr("boss_agent_cli.commands.crawl._launch_background_resume", lambda data_dir, run_id, **kwargs: launched.append(run_id))
	runner = CliRunner()
	_enable_research(runner, tmp_path)
	result = runner.invoke(cli, ["--data-dir", str(tmp_path), "--json", "crawl", "start", "AI", "--city", "杭州"])
	assert result.exit_code == 0, result.output
	run_id = json.loads(result.output)["data"]["run_id"]
	assert launched == [run_id]
	stopped = runner.invoke(cli, ["--data-dir", str(tmp_path), "--json", "crawl", "stop", run_id])
	assert stopped.exit_code == 0, stopped.output
	with CacheStore(tmp_path / "cache" / "boss_agent.db") as cache:
		assert cache.get_crawl_run(run_id)["stop_requested"] is True


def test_exporter_formats_workbook(tmp_path):
	paths = write_run_outputs(tmp_path, [{"title": "AI", "salary": "25K", "benefits": "双休", "post_description": "描述"}])
	workbook = load_workbook(paths["xlsx"])
	sheet = workbook.active
	assert sheet.column_dimensions["P"].width == 54
	assert sheet["P2"].alignment.wrap_text is not True
	assert sheet.row_dimensions[2].height == 20
